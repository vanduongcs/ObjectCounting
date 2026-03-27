"""
Test Edge Cases: Kiểm tra logic đếm và trạng thái UI.

Cách chạy: python -m tests.test_edge_cases
"""

import json
import sys
import os
import tempfile
from pathlib import Path
from unittest.mock import MagicMock

# Fix encoding cho Windows console
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr and hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

_REPO_ROOT = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

# Mock PyQt6 & ultralytics (không cần cài để test logic)
sys.modules['PyQt6'] = MagicMock()
sys.modules['PyQt6.QtWidgets'] = MagicMock()
sys.modules['PyQt6.QtCore'] = MagicMock()
sys.modules['PyQt6.QtGui'] = MagicMock()
sys.modules['ultralytics'] = MagicMock()

from src.services.counter_service import CounterService
from src.services import cache_service
from src.services import db_service
from src.services.tracklet_stitcher import TrackletStitcher
from src.services.timestamp_ocr import TimestampOCRService
from src.utils.export_excel_processor import export_to_excel
from src.utils.source_utils import (
    build_source_key,
    get_source_region,
    is_stream_source,
    load_source_region_store,
    save_source_region,
)
from src.utils.ui_state import compute_main_window_button_states
import configs.settings as settings

from openpyxl import load_workbook


# ======================================================================
# HELPER
# ======================================================================
_SOURCE_DIR = os.path.join(_REPO_ROOT, 'src')


def _read_source(relative_path):
    with open(os.path.join(_SOURCE_DIR, relative_path), 'r', encoding='utf-8') as f:
        return f.read()


def _read_repo_file(*parts):
    with open(os.path.join(_REPO_ROOT, *parts), 'r', encoding='utf-8') as f:
        return f.read()


def _make_det(id, label=0, cx=100, cy=100, w=50, h=50, conf=0.8):
    """Helper tạo detection dict cho tests."""
    return {"id": id, "label": label, "center": (cx, cy),
            "bbox_wh": (w, h), "conf": conf}


def _repeat_update(counter, detections, n=5):
    """Gửi cùng detections n lần liên tiếp để thoả debounce."""
    for _ in range(n):
        counter.update(detections)


# ======================================================================
# COUNTER SERVICE TESTS
# ======================================================================

def test_counter_basic_nhap():
    """XUAT → NHAP = đếm 1 NHẬP."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    nhap, xuat = counter.get_counts()
    assert nhap.get(0, 0) == 1
    assert xuat.get(0, 0) == 0
    print("✅ TEST 1 PASSED: Đếm NHẬP cơ bản")


def test_counter_basic_xuat():
    """NHAP → XUAT = đếm 1 XUẤT."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    nhap, xuat = counter.get_counts()
    assert xuat.get(0, 0) == 1
    assert nhap.get(0, 0) == 0
    print("✅ TEST 2 PASSED: Đếm XUẤT cơ bản")


def test_counter_near_line_still_counts():
    """Vật gần vạch vẫn đếm (không có buffer zone nữa)."""
    counter = CounterService((200, 0), (200, 400))
    # Bắt đầu ở x=195 (rất gần vạch, bên NHAP)
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (195, 200)}])
    # Di chuyển sang x=205 (rất gần vạch, bên XUAT)
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (205, 200)}])
    nhap, xuat = counter.get_counts()
    assert sum(nhap.values()) + sum(xuat.values()) == 1
    print("✅ TEST 3 PASSED: Gần vạch vẫn đếm đúng")


def test_counter_multiple_objects():
    """Nhiều vật cùng qua vạch."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [
        {"id": 1, "label": 0, "center": (150, 100)},
        {"id": 2, "label": 0, "center": (150, 200)},
        {"id": 3, "label": 1, "center": (150, 300)},
    ])
    _repeat_update(counter, [
        {"id": 1, "label": 0, "center": (250, 100)},
        {"id": 2, "label": 0, "center": (250, 200)},
        {"id": 3, "label": 1, "center": (250, 300)},
    ])
    nhap, xuat = counter.get_counts()
    assert xuat.get(0, 0) == 2
    assert xuat.get(1, 0) == 1
    print("✅ TEST 4 PASSED: Đếm nhiều vật cùng lúc")


def test_counter_no_double_count():
    """Qua vạch 1 lần chỉ đếm 1."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (300, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (350, 200)}])
    nhap, xuat = counter.get_counts()
    assert xuat.get(0, 0) == 1
    print("✅ TEST 5 PASSED: Không đếm trùng")


def test_counter_object_returns():
    """NHAP → XUAT → NHAP => bộ đếm net quay về 0 sau khi triệt tiêu."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    # Chờ cooldown hết
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}],
                   n=settings.COUNTER_COOLDOWN_FRAMES + 1)
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    nhap, xuat = counter.get_counts()
    assert xuat.get(0, 0) == 0
    assert nhap.get(0, 0) == 0
    assert len(counter.get_event_log()) == 2
    print("✅ TEST 6 PASSED: Vật quay lại triệt tiêu bộ đếm net")


def test_counter_net_counts_cancel_xuat():
    """XUAT → XUAT → NHAP => xuat giảm dần, không giữ cả hai phía cùng lúc."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}],
                   n=settings.COUNTER_COOLDOWN_FRAMES + 1)
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (250, 200)}],
                   n=settings.COUNTER_COOLDOWN_FRAMES + 1)
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (150, 200)}])
    nhap, xuat = counter.get_counts()
    assert nhap.get(0, 0) == 0
    assert xuat.get(0, 0) == 1
    print("✅ TEST 6b PASSED: NHẬP triệt tiêu XUẤT tồn trước đó")


def test_counter_net_counts_cancel_nhap():
    """Kịch bản user nêu: NHAP → NHAP → XUAT => còn 1 NHAP."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}],
                   n=settings.COUNTER_COOLDOWN_FRAMES + 1)
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (250, 200)}])
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (150, 200)}])
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (150, 200)}],
                   n=settings.COUNTER_COOLDOWN_FRAMES + 1)
    _repeat_update(counter, [{"id": 2, "label": 0, "center": (250, 200)}])
    nhap, xuat = counter.get_counts()
    assert nhap.get(0, 0) == 1
    assert xuat.get(0, 0) == 0
    print("✅ TEST 6c PASSED: XUẤT triệt tiêu NHẬP tồn trước đó")


def test_counter_memory_limit():
    """object_states phải được giới hạn."""
    counter = CounterService((200, 0), (200, 400))
    for i in range(settings.MAX_TRACKED_OBJECTS + 500):
        counter.update([{"id": i, "label": 0, "center": (150, 200)}])
    assert len(counter.object_states) <= settings.MAX_TRACKED_OBJECTS + 1
    print(f"✅ TEST 7 PASSED: Memory limit ({len(counter.object_states)} entries)")


def test_counter_diagonal_line():
    """Vạch chéo vẫn đếm đúng."""
    counter = CounterService((0, 0), (400, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (100, 300)}])
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (300, 100)}])
    nhap, xuat = counter.get_counts()
    assert sum(nhap.values()) + sum(xuat.values()) == 1
    print("✅ TEST 8 PASSED: Vạch chéo đếm đúng")


def test_counter_flicker_no_overcount():
    """Vật nhấp nháy qua vạch → chỉ đếm ít (anti-flicker)."""
    counter = CounterService((200, 0), (200, 400))
    _repeat_update(counter, [{"id": 1, "label": 0, "center": (150, 200)}])
    # Nhấp nháy: qua lại vạch mỗi frame (1 frame mỗi bên, < debounce=2)
    for _ in range(20):
        counter.update([{"id": 1, "label": 0, "center": (250, 200)}])
        counter.update([{"id": 1, "label": 0, "center": (150, 200)}])
    nhap, xuat = counter.get_counts()
    total = sum(nhap.values()) + sum(xuat.values())
    assert total <= 1, f"Flicker: expected ≤1, got {total}"
    print(f"✅ TEST 8b PASSED: Anti-flicker (total={total})")


def test_cache_finish_recording_uses_final_frame_count():
    """finish_recording phải lấy frame_count sau khi writer thread flush xong."""
    original_state = (
        cache_service._write_queue,
        cache_service._writer_thread,
        cache_service._video_writer,
        cache_service._output_path,
        cache_service._frame_count,
    )

    class _FakeQueue:
        def __init__(self):
            self.items = []

        def put(self, item):
            self.items.append(item)

    class _FakeThread:
        def join(self, timeout=None):
            cache_service._frame_count = 5

    class _FakeWriter:
        def __init__(self):
            self.released = False

        def release(self):
            self.released = True

    fd, output_path = tempfile.mkstemp(suffix=".mp4")
    os.close(fd)

    fake_queue = _FakeQueue()
    fake_thread = _FakeThread()
    fake_writer = _FakeWriter()

    try:
        cache_service._write_queue = fake_queue
        cache_service._writer_thread = fake_thread
        cache_service._video_writer = fake_writer
        cache_service._output_path = output_path
        cache_service._frame_count = 0

        path, count = cache_service.finish_recording()

        assert path == output_path
        assert count == 5
        assert fake_queue.items == [None]
        assert fake_writer.released is True
        assert os.path.exists(output_path)
        print("✅ TEST 8c PASSED: finish_recording lấy count sau khi flush")
    finally:
        (
            cache_service._write_queue,
            cache_service._writer_thread,
            cache_service._video_writer,
            cache_service._output_path,
            cache_service._frame_count,
        ) = original_state
        if os.path.exists(output_path):
            os.remove(output_path)


def test_export_allows_event_log_without_net_counts():
    """Vẫn phải export được nếu count net = 0 nhưng event log còn dữ liệu."""
    fd, output_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    os.remove(output_path)
    try:
        ok, _msg = export_to_excel(
            output_path,
            {},
            {},
            [{"label": "can_nho", "action": "Nhập", "timestamp": "00:00:05"}],
        )
        assert ok is True
        wb = load_workbook(output_path)
        try:
            assert "Tong hop" in wb.sheetnames
            assert "can_nho" in wb.sheetnames
            ws = wb["can_nho"]
            assert ws["A3"].value == "00:00:05"
        finally:
            wb.close()
        print("✅ TEST 8d PASSED: Export vẫn có sheet từ event log")
    finally:
        if os.path.exists(output_path):
            os.remove(output_path)


def test_stream_helper_supports_all_supported_schemes():
    """Nhận diện stream phải nhất quán cho http/https/rtsp/rtmp."""
    assert is_stream_source("http://camera.local/live") is True
    assert is_stream_source("https://camera.local/live") is True
    assert is_stream_source("rtsp://camera.local/live") is True
    assert is_stream_source("rtmp://camera.local/live") is True
    assert is_stream_source("D:\\video.mp4") is False
    print("✅ TEST 8e PASSED: Stream scheme helper đúng")


def test_source_key_normalizes_streams_and_paths():
    """Source key phải ổn định để line/ROI không bị lẫn giữa các nguồn."""
    stream_key = build_source_key(" HTTP://Camera.Local/Live ")
    stream_variant_key = build_source_key(" HTTP://Camera.Local/Live ", variant="rot=90")
    file_key = build_source_key(os.path.join("tests", "..", "demo.mp4"))
    assert stream_key == "http://camera.local/live"
    assert stream_variant_key == "http://camera.local/live::variant=rot=90"
    assert file_key == os.path.normcase(os.path.abspath(os.path.join("tests", "..", "demo.mp4")))
    print("✅ TEST 8e2 PASSED: Source key normalize đúng")


def test_source_region_store_roundtrip_is_per_source():
    """Line/ROI lưu theo source phải round-trip đúng và không rò sang source khác."""
    with tempfile.TemporaryDirectory() as tmpdir:
        store_path = Path(tmpdir) / "regions.json"
        store = {}
        rel = (0.1, 0.2, 0.3, 0.4)

        assert save_source_region(store_path, store, "video_a.mp4", rel, variant="rot=0") is True

        loaded = load_source_region_store(store_path)
        assert get_source_region(loaded, "video_a.mp4", variant="rot=0") == rel
        assert get_source_region(loaded, "video_a.mp4", variant="rot=90") is None
        assert get_source_region(loaded, "video_b.mp4") is None
    print("✅ TEST 8e3 PASSED: Source store per-source đúng")


def test_source_region_store_falls_back_to_last_used_when_requested():
    """Source mới có thể kế thừa line/ROI đang dùng nếu chưa có giá trị riêng."""
    with tempfile.TemporaryDirectory() as tmpdir:
        store_path = Path(tmpdir) / "regions.json"
        store = {}
        rel = (0.1, 0.2, 0.3, 0.4)

        assert save_source_region(store_path, store, "video_a.mp4", rel, variant="rot=0") is True

        loaded = load_source_region_store(store_path)
        assert get_source_region(
            loaded,
            "video_b.mp4",
            variant="rot=0",
            fallback_to_last_used=True,
        ) == rel
        assert get_source_region(
            loaded,
            "video_b.mp4",
            variant="rot=90",
            fallback_to_last_used=True,
        ) is None
    print("✅ TEST 8e3b PASSED: Fallback last-used đúng")


def test_source_region_store_ignores_legacy_global_rel():
    """File rel kiểu cũ không được tự apply cho source mới bất kỳ."""
    with tempfile.TemporaryDirectory() as tmpdir:
        store_path = Path(tmpdir) / "legacy.json"
        store_path.write_text(json.dumps({"rel": [0.1, 0.2, 0.3, 0.4]}), encoding="utf-8")
        loaded = load_source_region_store(store_path)
        assert get_source_region(loaded, "video_a.mp4") is None
    print("✅ TEST 8e4 PASSED: Legacy global rel không auto-apply")


def test_source_region_store_can_upgrade_legacy_rel_via_fallback():
    """File rel kiểu cũ vẫn dùng làm fallback nếu caller yêu cầu kế thừa."""
    with tempfile.TemporaryDirectory() as tmpdir:
        store_path = Path(tmpdir) / "legacy.json"
        rel = (0.1, 0.2, 0.3, 0.4)
        store_path.write_text(json.dumps({"rel": list(rel)}), encoding="utf-8")
        loaded = load_source_region_store(store_path)
        assert get_source_region(
            loaded,
            "video_a.mp4",
            fallback_to_last_used=True,
        ) == rel
    print("✅ TEST 8e4b PASSED: Legacy rel fallback đúng")


def test_db_json_loader_tolerates_invalid_json():
    """JSON hỏng phải rơi về default thay vì làm sập panel history."""
    assert db_service._loads_json_or_default("{bad", {}) == {}
    assert db_service._loads_json_or_default("[bad", []) == []
    assert db_service._loads_json_or_default('{"ok": 1}', {}) == {"ok": 1}
    print("✅ TEST 8f PASSED: DB tolerate invalid JSON")


def test_tracklet_stitcher_lost_tracks_are_capped():
    """Lost IDs tích tụ lâu không được tăng vô hạn."""
    stitcher = TrackletStitcher(fps=30)
    stitcher.min_observe_frames = 1
    stitcher.missing_to_lost = 1
    stitcher.max_lost_tracks = 5
    stitcher.max_lost_frames = 999

    for idx in range(20):
        stitcher.process([{"id": idx, "label": 0, "center": (50.0, 50.0), "bbox_wh": (10.0, 10.0), "conf": 0.9}])
        stitcher.process([])

    assert len(stitcher._lost) <= 5
    print("✅ TEST 8g PASSED: Lost tracks được cap")


def test_timestamp_ocr_normalizes_frame_path_to_video_time():
    """Frame index encoded in crop filename phải đổi về HH:MM:SS."""
    service = TimestampOCRService()
    normalized = service.normalize_event_log(
        [{"label": "can_nho", "action": "Nhập", "timestamp": "video_00000050.jpg"}],
        fps=25,
    )
    assert normalized[0]["timestamp"] == "00:00:02"
    print("✅ TEST 8h PASSED: Timestamp OCR normalize frame path đúng")


def test_timestamp_ocr_finalize_keeps_valid_timestamp_text():
    """Timestamp hợp lệ sẵn có không được bị đổi khi finalize."""
    service = TimestampOCRService()
    finalized = service.finalize_event_log(
        [{"label": "can_nho", "action": "Nhập", "timestamp": "19-03-2026 08:07:06"}],
        fps=25,
    )
    assert finalized[0]["timestamp"] == "19-03-2026 08:07:06"
    print("✅ TEST 8i PASSED: Timestamp OCR giữ nguyên timestamp hợp lệ")


def test_timestamp_ocr_picks_first_valid_output_timestamp():
    """Timestamp đầu tiên không hợp lệ phải bị bỏ qua, lấy timestamp hợp lệ kế tiếp."""
    service = TimestampOCRService()
    parsed_dt = service.first_valid_event_timestamp([
        {"timestamp": "32-13-2026 99:99:99"},
        {"timestamp": "19-03-2026 08:07:06"},
        {"timestamp": "20-03-2026 09:10:11"},
    ])
    assert parsed_dt is not None
    assert service.build_output_video_name(parsed_dt) == "08 giờ 07 phút_ngày 19 tháng 03 năm 2026.mp4"
    print("✅ TEST 8j PASSED: Lấy đúng timestamp hợp lệ đầu tiên để đặt tên file")


def test_timestamp_ocr_builds_excel_name_from_events():
    """Tên file Excel phải dùng cùng timestamp logic như video output."""
    file_name = TimestampOCRService.build_output_file_name_from_events(
        [
            {"timestamp": "invalid"},
            {"timestamp": "19-03-2026 08:07:06"},
        ],
        "xlsx",
    )
    assert file_name == "08 giờ 07 phút_ngày 19 tháng 03 năm 2026.xlsx"
    print("✅ TEST 8j2 PASSED: Tạo tên Excel từ event log đúng")


def test_cache_rename_output_video_uses_timestamp_file_name():
    """Video output phải đổi tên sang format ngày giờ mới sau khi xử lý."""
    with tempfile.TemporaryDirectory() as tmpdir:
        original_path = os.path.join(tmpdir, "raw_output.mp4")
        with open(original_path, "wb") as handle:
            handle.write(b"test")

        renamed_path = cache_service.rename_output_video(
            original_path,
            "08 giờ 07 phút_ngày 19 tháng 03 năm 2026.mp4",
        )

        assert os.path.exists(renamed_path)
        assert not os.path.exists(original_path)
        assert os.path.basename(renamed_path) == "08 giờ 07 phút_ngày 19 tháng 03 năm 2026.mp4"
    print("✅ TEST 8k PASSED: Rename video output theo timestamp hoạt động")


# ======================================================================
# TRACKLET STITCHER TESTS -- tạm bỏ (stitcher disabled)
# ======================================================================


# ======================================================================
# UI STATE TESTS (kiểm tra source code)
# ======================================================================

def test_ui_running_state_toggle():
    """Idle + có source + có line => bật đúng các nút cần dùng."""
    states = compute_main_window_button_states(
        source_loaded=True,
        running=False,
        has_line=True,
        is_stream=False,
    )
    assert states["choose"] is True
    assert states["camera"] is True
    assert states["draw"] is True
    assert states["rotate"] is True
    assert states["timestamp"] is True
    assert states["start"] is True
    assert states["pause"] is False
    assert states["stop"] is False
    print("✅ TEST 15 PASSED: Button state idle/video đúng")


def test_ui_button_state_requires_line():
    """Không có line thì không cho Start."""
    states = compute_main_window_button_states(
        source_loaded=True,
        running=False,
        has_line=False,
        is_stream=False,
    )
    assert states["start"] is False
    assert states["draw"] is True
    print("✅ TEST 15b PASSED: Start bị khóa khi chưa có line")


def test_ui_button_state_stream_pause_disabled():
    """Camera/stream đang chạy thì Pause bị disable."""
    states = compute_main_window_button_states(
        source_loaded=True,
        running=True,
        has_line=True,
        is_stream=True,
    )
    assert states["pause"] is False
    assert states["stop"] is True
    assert states["start"] is False
    print("✅ TEST 15c PASSED: Stream running không bật Pause")


def test_close_event_has_cleanup():
    """closeEvent phải gọi _cleanup_threads."""
    source = _read_source('views/main_window.py')
    lines = source.split('\n')
    found_close = False
    for i, line in enumerate(lines):
        if 'def closeEvent' in line:
            found_close = True
        if found_close and '_cleanup_threads' in line:
            break
    assert found_close, "closeEvent không tìm thấy"
    assert '_cleanup_threads' in source
    print("✅ TEST 16 PASSED: closeEvent cleanup đúng")


def test_start_cleans_old_threads():
    """start_processing phải cleanup thread cũ trước khi tạo mới."""
    source = _read_source('views/main_window.py')
    cleanup_pos = source.index('_cleanup_threads')
    thread_pos = source.index('start_extraction')
    assert cleanup_pos < thread_pos
    print("✅ TEST 17 PASSED: start_processing cleanup thread cũ đúng")


def test_start_without_video_returns():
    """start_processing phải kiểm tra current_video_path."""
    source = _read_source('views/main_window.py')
    assert 'current_video_path' in source
    print("✅ TEST 18 PASSED: start guard check OK")


def test_frame_extractor_reconnect_limit():
    """FrameExtractor phải có giới hạn reconnect."""
    source = _read_source('services/frame_extractor.py')
    assert 'MAX_STREAM_FAILURES' in source
    assert 'consecutive_failures' in source
    print("✅ TEST 19 PASSED: Stream reconnect limit đúng")


def test_frame_extractor_open_fail_unblocks_ai():
    """Open fail phải signal FPS-ready và sentinel để AI thoát nhanh."""
    source = _read_source('services/frame_extractor.py')
    start = source.index('if not cap.isOpened():')
    end = source.index('video_fps = self._get_video_fps')
    block = source[start:end]
    assert '_fps_ready.set()' in block
    assert '_enqueue_frame(None)' in block
    print("✅ TEST 19b PASSED: Open fail unblocks AI đúng")


def test_virtual_line_scaling():
    """start_processing phải scale vạch ảo."""
    source = _read_source('views/main_window.py')
    assert '_scale_line' in source
    print("✅ TEST 20 PASSED: Virtual line scaling logic OK")


def test_virtual_line_persistence_hooks():
    """MainWindow phải lưu/load/restore vạch ảo từ file."""
    source = _read_source('views/main_window.py')
    assert '_save_virtual_line' in source
    assert '_load_virtual_line' in source
    assert '_restore_virtual_line_for_current_frame' in source
    assert 'VIRTUAL_LINE_PATH' in source
    print("✅ TEST 21 PASSED: Virtual line persistence hooks OK")


def test_main_window_uses_source_scoped_persistence():
    """MainWindow phải dùng helper per-source thay vì rel global duy nhất."""
    source = _read_source('views/main_window.py')
    assert 'load_source_region_store' in source
    assert 'save_source_region' in source
    assert '_current_persistence_variant' in source
    assert 'rot=' in source
    assert 'variant=self._current_persistence_variant()' in source
    assert '_restore_timestamp_space_for_current_source' in source
    print("✅ TEST 21b PASSED: MainWindow dùng source-scoped persistence")


def test_main_window_draws_preview_overlays_only_when_idle():
    """Đang xử lý thì preview không nên mutate frame thêm lần nữa."""
    source = _read_source('views/main_window.py')
    assert 'render_preview_overlays = not self._is_processing_active()' in source
    print("✅ TEST 21c PASSED: Preview overlay chỉ vẽ khi idle")


def test_ai_service_deduplicates_count_emits():
    """Count signal phải được dedupe để giảm rebuild UI thừa."""
    source = _read_source('services/ai_service.py')
    assert '_last_emitted_counts' in source
    assert 'if counts_key == self._last_emitted_counts:' in source
    print("✅ TEST 21d PASSED: AI count emit dedupe đúng")


def test_ai_service_compensates_pause_time_in_sync_clock():
    """Pause video file không được làm playback sync nhảy cóc frame."""
    source = _read_source('services/ai_service.py')
    assert 'def _wait_while_paused' in source
    assert 'start_time += time.time() - pause_started_at' in source
    print("✅ TEST 21e PASSED: AI pause compensation đúng")


def test_ai_service_supports_startup_preload():
    """AIService phải có hook preload detector để warmup trước khi show UI."""
    source = _read_source('services/ai_service.py')
    assert 'def preload_detector' in source
    assert '_reuse_preloaded_detector = True' in source
    print("✅ TEST 21f PASSED: AIService có preload detector")


def test_main_preloads_detector_before_show():
    """App phải warmup detector trước khi hiện MainWindow."""
    source = _read_source('main.py')
    assert 'ai_service.preload_detector()' in source
    assert source.index('ai_service.preload_detector()') < source.index('window.show()')
    print("✅ TEST 21g PASSED: main preload detector trước khi show UI")


def test_ai_service_has_separate_box_conf_toggle():
    """AIService phải có flag riêng để bật/tắt conf của bounding box."""
    source = _read_source('services/ai_service.py')
    assert 'self.show_box_conf = False' in source
    assert 'show_conf=self.show_box_conf' in source
    print("✅ TEST 21h PASSED: AIService có box conf toggle riêng")


def test_main_window_has_conf_switch():
    """MainWindow phải có switch UI để bật/tắt hiển thị conf."""
    source = _read_source('views/main_window.py')
    assert 'Hiển thị conf' in source
    assert 'self.show_conf_switch' in source
    assert 'def toggle_show_box_conf' in source
    print("✅ TEST 21i PASSED: MainWindow có switch conf")


def test_detector_render_supports_show_conf():
    """Detector render phải cho phép ẩn riêng conf mà vẫn giữ box."""
    source = _read_source('services/detector.py')
    assert 'def render(self, frame, raw_results, show_boxes=True, show_conf=True)' in source
    assert 'def _draw_prediction_rows' in source
    assert 'return f"{name} {score:.1f}"' in source
    print("✅ TEST 21j PASSED: Detector render hỗ trợ show_conf")


def test_settings_interface_file_exists():
    """UI config phải có file riêng để gom layout và sizing tuning."""
    source = _read_repo_file('configs', 'settings_interface.py')
    assert 'SIDE_PANEL_WIDTH' in source
    assert 'HISTORY_DRAWER_EXPANDED_WIDTH' in source
    print("✅ TEST 21k PASSED: settings_interface.py tồn tại")


def test_settings_theme_file_exists():
    """Theme config phải có file riêng để gom màu và stylesheet."""
    source = _read_repo_file('configs', 'settings_theme.py')
    assert 'MAIN_WINDOW_STYLESHEET' in source
    assert 'TOGGLE_TRACK_ON_COLOR' in source
    assert 'CLASS_COLOR_PALETTE' in source
    print("✅ TEST 21k2 PASSED: settings_theme.py tồn tại")


def test_main_window_uses_settings_interface():
    """MainWindow phải đọc layout từ settings_interface."""
    source = _read_source('views/main_window.py')
    assert 'import configs.settings_interface as ui_settings' in source
    assert 'ui_settings.SIDE_PANEL_WIDTH' in source
    print("✅ TEST 21l PASSED: MainWindow dùng settings_interface")


def test_main_window_uses_settings_theme():
    """MainWindow phải đọc stylesheet màu từ settings_theme."""
    source = _read_source('views/main_window.py')
    assert 'import configs.settings_theme as theme_settings' in source
    assert 'theme_settings.MAIN_WINDOW_STYLESHEET' in source
    print("✅ TEST 21l2 PASSED: MainWindow dùng settings_theme")


def test_visual_helpers_use_split_interface_and_theme():
    """Widget phụ và helper vẽ phải tách layout và màu ra 2 config riêng."""
    detector_source = _read_source('services/detector.py')
    draw_line_source = _read_source('utils/draw_line.py')
    toggle_source = _read_source('views/toggle_switch.py')
    assert 'import configs.settings_interface as ui_settings' in detector_source
    assert 'import configs.settings_theme as theme_settings' in detector_source
    assert 'import configs.settings_interface as ui_settings' in draw_line_source
    assert 'import configs.settings_theme as theme_settings' in draw_line_source
    assert 'import configs.settings_interface as ui_settings' in toggle_source
    assert 'import configs.settings_theme as theme_settings' in toggle_source
    print("✅ TEST 21m PASSED: Helper UI tách settings_interface/settings_theme")


def test_main_window_persists_window_state():
    """MainWindow phai luu/khoi phuc geometry va maximize mac dinh lan dau."""
    settings_source = _read_repo_file('configs', 'settings.py')
    interface_source = _read_repo_file('configs', 'settings_interface.py')
    source = _read_source('views/main_window.py')
    assert 'WINDOW_STATE_PATH' in settings_source
    assert 'WINDOW_START_MAXIMIZED = True' in interface_source
    assert 'def _load_window_state' in source
    assert 'def _save_window_state' in source
    assert 'def _restore_window_state' in source
    assert 'self._restore_window_state()' in source
    assert 'self._save_window_state()' in source
    assert 'Qt.WindowState.WindowMaximized' in source
    print("TEST 21n PASSED: MainWindow persist window state dung")


# ======================================================================
# RUNNER
# ======================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("CHAY EDGE CASE TESTS")
    print("=" * 60)

    tests = [
        ("Dem NHAP co ban", test_counter_basic_nhap),
        ("Dem XUAT co ban", test_counter_basic_xuat),
        ("Gan vach van dem", test_counter_near_line_still_counts),
        ("Dem nhieu vat cung luc", test_counter_multiple_objects),
        ("Khong dem trung", test_counter_no_double_count),
        ("Vat quay lai", test_counter_object_returns),
        ("Triet tieu XUAT bang NHAP", test_counter_net_counts_cancel_xuat),
        ("Triet tieu NHAP bang XUAT", test_counter_net_counts_cancel_nhap),
        ("Memory limit", test_counter_memory_limit),
        ("Vach cheo", test_counter_diagonal_line),
        ("Anti-flicker", test_counter_flicker_no_overcount),
        ("Cache finish flush count", test_cache_finish_recording_uses_final_frame_count),
        ("Export event-only", test_export_allows_event_log_without_net_counts),
        ("Stream helper schemes", test_stream_helper_supports_all_supported_schemes),
        ("Source key normalize", test_source_key_normalizes_streams_and_paths),
        ("Source store per-source", test_source_region_store_roundtrip_is_per_source),
        ("Source store fallback last-used", test_source_region_store_falls_back_to_last_used_when_requested),
        ("Legacy rel isolated", test_source_region_store_ignores_legacy_global_rel),
        ("Legacy rel fallback", test_source_region_store_can_upgrade_legacy_rel_via_fallback),
        ("DB invalid JSON tolerance", test_db_json_loader_tolerates_invalid_json),
        ("Tracklet lost cap", test_tracklet_stitcher_lost_tracks_are_capped),
        ("Timestamp normalize frame path", test_timestamp_ocr_normalizes_frame_path_to_video_time),
        ("Timestamp keep valid text", test_timestamp_ocr_finalize_keeps_valid_timestamp_text),
        ("Timestamp pick first valid", test_timestamp_ocr_picks_first_valid_output_timestamp),
        ("Timestamp build excel name", test_timestamp_ocr_builds_excel_name_from_events),
        ("Cache rename output by timestamp", test_cache_rename_output_video_uses_timestamp_file_name),

        ("UI running state toggle", test_ui_running_state_toggle),
        ("UI requires line", test_ui_button_state_requires_line),
        ("UI stream pause disabled", test_ui_button_state_stream_pause_disabled),
        ("closeEvent cleanup", test_close_event_has_cleanup),
        ("start cleanup thread cu", test_start_cleans_old_threads),
        ("start guard check", test_start_without_video_returns),
        ("Stream reconnect limit", test_frame_extractor_reconnect_limit),
        ("Open fail unblocks AI", test_frame_extractor_open_fail_unblocks_ai),
        ("Virtual line scaling", test_virtual_line_scaling),
        ("Virtual line persistence hooks", test_virtual_line_persistence_hooks),
        ("MainWindow source persistence", test_main_window_uses_source_scoped_persistence),
        ("MainWindow idle preview overlays", test_main_window_draws_preview_overlays_only_when_idle),
        ("AI count emit dedupe", test_ai_service_deduplicates_count_emits),
        ("AI pause compensation", test_ai_service_compensates_pause_time_in_sync_clock),
        ("AI startup preload", test_ai_service_supports_startup_preload),
        ("Main preload before show", test_main_preloads_detector_before_show),
        ("AI box conf toggle", test_ai_service_has_separate_box_conf_toggle),
        ("Main conf switch", test_main_window_has_conf_switch),
        ("Detector show_conf render", test_detector_render_supports_show_conf),
        ("Settings interface exists", test_settings_interface_file_exists),
        ("Settings theme exists", test_settings_theme_file_exists),
        ("MainWindow uses settings interface", test_main_window_uses_settings_interface),
        ("MainWindow uses settings theme", test_main_window_uses_settings_theme),
        ("Visual helpers split interface and theme", test_visual_helpers_use_split_interface_and_theme),
        ("MainWindow persists window state", test_main_window_persists_window_state),
    ]

    passed = failed = 0
    for name, test_fn in tests:
        try:
            test_fn()
            passed += 1
        except Exception as e:
            print(f"FAILED: {name}")
            print(f"   Error: {e}")
            failed += 1

    print()
    print("=" * 60)
    print(f"KET QUA: {passed}/{passed + failed} passed, {failed} failed")
    if failed == 0:
        print("TAT CA TESTS DEU PASS!")
    else:
        print(f"CO {failed} TEST THAT BAI!")
    print("=" * 60)
