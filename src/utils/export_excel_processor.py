"""
Export Excel helpers.

Workbook layout:
- Tong hop: one row per label
- <label>: two side-by-side sections for Nhap and Xuat events
"""

import os
import re
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

import configs.settings as settings

_INVALID_SHEET_CHARS_RE = re.compile(r"[\[\]:*?/\\]")


def _normalize_counts(counts):
    if not isinstance(counts, dict):
        return {}
    normalized = {}
    for label, value in counts.items():
        key = str(label)
        try:
            normalized[key] = int(value)
        except Exception:
            try:
                normalized[key] = int(float(value))
            except Exception:
                normalized[key] = 0
    return normalized


def _normalize_event_log(event_log):
    if not isinstance(event_log, list):
        return []
    normalized = []
    for evt in event_log:
        if not isinstance(evt, dict):
            continue
        normalized.append({
            "label": str(evt.get("label", "")),
            "action": str(evt.get("action", "")),
            "timestamp": evt.get("timestamp", ""),
        })
    return normalized


def _make_sheet_name(label, used_names):
    base = _INVALID_SHEET_CHARS_RE.sub("_", str(label)).strip() or "Unknown"
    base = base[:31]
    if base.lower() not in used_names:
        used_names.add(base.lower())
        return base

    idx = 2
    while True:
        suffix = f"_{idx}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate.lower() not in used_names:
            used_names.add(candidate.lower())
            return candidate
        idx += 1


def _extract_frame_index(path):
    base = os.path.basename(path or "")
    match = re.search(r"_(\d{8})\.jpg$", base)
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def _frame_index_to_video_time(frame_idx, fps=None):
    fps_value = float(fps) if fps and fps > 0 else float(getattr(settings, "DEFAULT_VIDEO_FPS", 25))
    if frame_idx is None or fps_value <= 0:
        return ""
    total_seconds = int(frame_idx / fps_value)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _normalize_export_path(filepath):
    path = os.path.abspath(str(filepath or "").strip())
    if not path:
        return ""
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    return path


def _build_fallback_export_path(filepath):
    base, ext = os.path.splitext(filepath)
    ext = ext or ".xlsx"
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = f"{base}_{stamp}{ext}"
    idx = 2
    while os.path.exists(candidate):
        candidate = f"{base}_{stamp}_{idx}{ext}"
        idx += 1
    return candidate


def _normalize_timestamp(value):
    if isinstance(value, str) and value and os.path.exists(value):
        return _frame_index_to_video_time(_extract_frame_index(value))
    return str(value) if value is not None else ""


def _split_events_by_action(event_log, label):
    nhap_rows = []
    xuat_rows = []
    for evt in event_log:
        if evt.get("label") != label:
            continue
        row = [evt.get("timestamp", "")]
        action = str(evt.get("action", "")).lower()
        if "nh" in action:
            nhap_rows.append(row)
        elif "xu" in action:
            xuat_rows.append(row)
    return nhap_rows, xuat_rows


def _collect_all_labels(count_nhap, count_xuat, event_log):
    labels = set(count_nhap.keys()) | set(count_xuat.keys())
    for evt in event_log:
        label = str(evt.get("label", "")).strip()
        if label:
            labels.add(label)
    return sorted(labels)


def _style_header_row(ws, row_idx, start_col, end_col, fill_color=None):
    header_font = Font(bold=True, size=11)
    thin_border = Border(bottom=Side(style="thin", color="999999"))
    fill = PatternFill("solid", fgColor=fill_color) if fill_color else None
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = header_font
        cell.border = thin_border
        if fill is not None:
            cell.fill = fill


def export_to_excel(filepath, count_nhap, count_xuat, event_log=None):
    """
    Export current net counts and event log into Excel.

    Returns:
        (success: bool, message: str)
    """
    tmp_path = ""
    try:
        filepath = _normalize_export_path(filepath)
        if not filepath:
            return False, "Loi xuat file: Duong dan file khong hop le."

        target_dir = os.path.dirname(filepath) or os.getcwd()
        os.makedirs(target_dir, exist_ok=True)

        count_nhap = _normalize_counts(count_nhap)
        count_xuat = _normalize_counts(count_xuat)
        event_log = _normalize_event_log(event_log)
        processed_events = []
        for evt in event_log:
            processed_events.append({
                "label": evt.get("label", ""),
                "action": evt.get("action", ""),
                "timestamp": _normalize_timestamp(evt.get("timestamp", "")),
            })
        all_labels = _collect_all_labels(count_nhap, count_xuat, processed_events)

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "Tong hop"
        used_sheet_names = {ws_summary.title.lower()}

        ws_summary.append(["Loai", "Nhap hien tai", "Xuat hien tai"])
        _style_header_row(ws_summary, 1, 1, 3, fill_color="D9EAF7")
        for label in all_labels:
            ws_summary.append([
                label,
                count_nhap.get(label, 0),
                count_xuat.get(label, 0),
            ])
        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 16
        ws_summary.column_dimensions["C"].width = 16

        for label in all_labels:
            sheet_name = _make_sheet_name(label, used_sheet_names)
            ws = wb.create_sheet(title=sheet_name)

            nhap_rows, xuat_rows = _split_events_by_action(processed_events, label)

            ws["A1"] = "Nhap"
            ws["C1"] = "Xuat"
            _style_header_row(ws, 1, 1, 1, fill_color="E2F0D9")
            _style_header_row(ws, 1, 3, 3, fill_color="FCE4D6")

            ws["A2"] = "Thoi gian"
            ws["C2"] = "Thoi gian"
            _style_header_row(ws, 2, 1, 1)
            _style_header_row(ws, 2, 3, 3)

            row_count = max(len(nhap_rows), len(xuat_rows), 1)
            for idx in range(row_count):
                if idx < len(nhap_rows):
                    ws.cell(row=idx + 3, column=1, value=nhap_rows[idx][0])
                if idx < len(xuat_rows):
                    ws.cell(row=idx + 3, column=3, value=xuat_rows[idx][0])

            ws.column_dimensions["A"].width = 22
            ws.column_dimensions["B"].width = 4
            ws.column_dimensions["C"].width = 22

        tmp_fd, tmp_path = tempfile.mkstemp(
            prefix="._excel_export_",
            suffix=".xlsx",
            dir=target_dir,
        )
        os.close(tmp_fd)
        wb.save(tmp_path)
        wb.close()

        try:
            os.replace(tmp_path, filepath)
            tmp_path = ""
            return True, f"Xuat file thanh cong:\n{filepath}"
        except PermissionError:
            fallback_path = _build_fallback_export_path(filepath)
            os.replace(tmp_path, fallback_path)
            tmp_path = ""
            return True, (
                "File dich dang duoc mo hoac dang bi khoa.\n"
                f"Da luu sang file khac:\n{fallback_path}"
            )

    except PermissionError:
        return False, (
            "Loi xuat file: Khong co quyen ghi hoac file dang bi ung dung khac khoa.\n"
            f"Duong dan: {filepath}"
        )
    except Exception as e:
        return False, f"Loi xuat file: {str(e)}"
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
